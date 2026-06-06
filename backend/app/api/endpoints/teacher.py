from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func
from sqlalchemy.orm import selectinload
import datetime
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    Workbook = None

from ...database import get_db
from ...models import (
    User,
    UserRole,
    StudyGroup,
    Assignment,
    Grade,
    Topic,
    Discipline,
    Message,
    MessageStatus,
    Attendance,
    Schedule,
    student_group_association,
)
from ...auth import get_current_user

router = APIRouter()


def _collect_course_ids(groups):
    course_ids = set()
    for group in groups:
        course_ids.update(c.id for c in group.courses)
    return list(course_ids)


def require_teacher_or_admin(current_user: User):
    if current_user.role not in [UserRole.teacher, UserRole.admin]:
        raise HTTPException(status_code=403, detail="Teacher or Admin access required")


@router.get("/stats")
async def get_teacher_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    require_teacher_or_admin(current_user)

    # For admins we return aggregate values; for teachers filter by teacher_id.
    teacher_filter = [] if current_user.role == UserRole.admin else [StudyGroup.teacher_id == current_user.id]

    groups_result = await db.execute(
        select(StudyGroup)
        .where(*teacher_filter)
        .options(selectinload(StudyGroup.courses))
    )
    groups = groups_result.scalars().all()

    group_ids = [g.id for g in groups]
    course_ids = _collect_course_ids(groups)

    total_groups = len(group_ids)

    if group_ids:
        students_result = await db.execute(
            select(func.count(func.distinct(User.id)))
            .join(StudyGroup.students)
            .where(StudyGroup.id.in_(group_ids), User.role == UserRole.student)
        )
        total_students = students_result.scalar() or 0
    else:
        total_students = 0

    # Assignments in courses taught by the teacher.
    if course_ids:
        assignments_result = await db.execute(
            select(Assignment)
            .join(Topic, Assignment.topic_id == Topic.id)
            .join(Discipline, Topic.discipline_id == Discipline.id)
            .where(Discipline.course_id.in_(course_ids))
        )
        assignments = assignments_result.scalars().all()
    else:
        assignments = []

    today = datetime.date.today()
    pending_assignments = sum(
        1 for a in assignments if (a.due_date is None or a.due_date >= today)
    )
    overdue_assignments = sum(
        1 for a in assignments if (a.due_date is not None and a.due_date < today)
    )

    # Pending questions scoped to this teacher: addressed to them, broadcast,
    # or sent by one of their students. Admins see all.
    pending_q = select(func.count(func.distinct(Message.id))).where(
        Message.is_escalated == True,
        Message.status == MessageStatus.pending,
    )
    if current_user.role != UserRole.admin:
        student_ids_sub = (
            select(student_group_association.c.student_id)
            .join(StudyGroup, StudyGroup.id == student_group_association.c.group_id)
            .where(StudyGroup.teacher_id == current_user.id)
        )
        pending_q = pending_q.where(
            (Message.receiver_id == current_user.id)
            | (Message.receiver_id.is_(None))
            | (Message.sender_id.in_(student_ids_sub))
        )
    pending_questions_result = await db.execute(pending_q)
    pending_questions = pending_questions_result.scalar() or 0

    # Attendance averaged only for this teacher's schedule entries.
    if current_user.role == UserRole.admin:
        attendance_result = await db.execute(select(Attendance.status))
    else:
        attendance_result = await db.execute(
            select(Attendance.status)
            .join(Schedule, Attendance.schedule_id == Schedule.id)
            .where(Schedule.teacher_id == current_user.id)
        )
    attendance_rows = attendance_result.scalars().all()
    
    # Count present attendances (status == "present")
    from ...models import AttendanceStatus
    present_count = sum(1 for s in attendance_rows if s == AttendanceStatus.present)
    total_count = len(attendance_rows)
    avg_attendance = round((present_count / total_count) * 100) if total_count > 0 else 0

    return {
        "total_students": total_students,
        "total_groups": total_groups,
        "pending_assignments": pending_assignments,
        "overdue_assignments": overdue_assignments,
        "pending_questions": pending_questions,
        "avg_attendance": avg_attendance,
    }


@router.get("/students")
async def get_teacher_students(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    require_teacher_or_admin(current_user)

    groups_query = select(StudyGroup).options(
        selectinload(StudyGroup.students),
        selectinload(StudyGroup.courses),
    )
    if current_user.role != UserRole.admin:
        groups_query = groups_query.where(StudyGroup.teacher_id == current_user.id)

    groups_result = await db.execute(groups_query)
    groups = groups_result.scalars().all()

    if not groups:
        return []

    group_ids = [g.id for g in groups]
    course_ids = _collect_course_ids(groups)

    # Build student -> one representative group name map and course ids for filtering.
    student_group_name = {}
    student_course_ids = {}
    for group in groups:
        group_course_ids = [c.id for c in group.courses]
        for student in group.students:
            if student.id not in student_group_name:
                student_group_name[student.id] = group.name
            if student.id not in student_course_ids:
                student_course_ids[student.id] = set()
            student_course_ids[student.id].update(group_course_ids)

    students_result = await db.execute(
        select(User)
        .join(StudyGroup.students)
        .where(StudyGroup.id.in_(group_ids), User.role == UserRole.student)
    )
    students = students_result.scalars().all()

    # Deduplicate by id
    dedup = {}
    for student in students:
        dedup[student.id] = student

    rows = []
    for student in dedup.values():
        grades_query = (
            select(Grade.score)
            .join(Assignment, Grade.assignment_id == Assignment.id)
            .join(Topic, Assignment.topic_id == Topic.id)
            .join(Discipline, Topic.discipline_id == Discipline.id)
            .where(Grade.student_id == student.id)
        )
        if course_ids:
            grades_query = grades_query.where(Discipline.course_id.in_(course_ids))

        scores_result = await db.execute(grades_query)
        scores = scores_result.scalars().all()
        avg_grade = round(sum(scores) / len(scores)) if scores else 0

        attendance_query = (
            select(Attendance.status)
            .join(Schedule, Attendance.schedule_id == Schedule.id)
            .where(Attendance.student_id == student.id)
        )
        if current_user.role != UserRole.admin:
            attendance_query = attendance_query.where(Schedule.teacher_id == current_user.id)

        attendance_result = await db.execute(attendance_query)
        attendance_rows = attendance_result.scalars().all()
        
        # Count present attendances
        from ...models import AttendanceStatus
        present_count = sum(1 for s in attendance_rows if s == AttendanceStatus.present)
        attendance = round((present_count / len(attendance_rows)) * 100) if attendance_rows else 0

        rows.append(
            {
                "id": student.id,
                "full_name": student.full_name,
                "email": student.email,
                "group": student_group_name.get(student.id, "—"),
                "course_ids": sorted(list(student_course_ids.get(student.id, set()))),
                "telegram_linked": bool(student.telegram_id),
                "avg_grade": avg_grade,
                "attendance": attendance,
            }
        )

    rows.sort(key=lambda x: x["full_name"])
    return rows


@router.get("/attendance-report/export")
async def export_attendance_report(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export attendance report for teacher's students as XLSX file"""
    require_teacher_or_admin(current_user)

    # Get students (similar logic to get_teacher_students)
    groups_query = select(StudyGroup).options(
        selectinload(StudyGroup.students),
        selectinload(StudyGroup.courses),
    )
    if current_user.role != UserRole.admin:
        groups_query = groups_query.where(StudyGroup.teacher_id == current_user.id)

    groups_result = await db.execute(groups_query)
    groups = groups_result.scalars().all()

    if not groups:
        rows = []
    else:
        group_ids = [g.id for g in groups]
        course_ids = _collect_course_ids(groups)

        # Build student -> one representative group name map.
        student_group_name = {}
        for group in groups:
            for student in group.students:
                if student.id not in student_group_name:
                    student_group_name[student.id] = group.name

        students_result = await db.execute(
            select(User)
            .join(StudyGroup.students)
            .where(StudyGroup.id.in_(group_ids), User.role == UserRole.student)
        )
        students = students_result.scalars().all()

        # Deduplicate by id
        dedup = {}
        for student in students:
            dedup[student.id] = student

        rows = []
        for student in dedup.values():
            grades_query = (
                select(Grade.score)
                .join(Assignment, Grade.assignment_id == Assignment.id)
                .join(Topic, Assignment.topic_id == Topic.id)
                .join(Discipline, Topic.discipline_id == Discipline.id)
                .where(Grade.student_id == student.id)
            )
            if course_ids:
                grades_query = grades_query.where(Discipline.course_id.in_(course_ids))

            scores_result = await db.execute(grades_query)
            scores = scores_result.scalars().all()
            avg_grade = round(sum(scores) / len(scores)) if scores else 0

            attendance_query = (
                select(Attendance.status)
                .join(Schedule, Attendance.schedule_id == Schedule.id)
                .where(Attendance.student_id == student.id)
            )
            if current_user.role != UserRole.admin:
                attendance_query = attendance_query.where(Schedule.teacher_id == current_user.id)

            attendance_result = await db.execute(attendance_query)
            attendance_rows = attendance_result.scalars().all()
            
            # Count present attendances
            from ...models import AttendanceStatus
            present_count = sum(1 for s in attendance_rows if s == AttendanceStatus.present)
            attendance = round((present_count / len(attendance_rows)) * 100) if attendance_rows else 0

            rows.append(
                {
                    "id": student.id,
                    "full_name": student.full_name,
                    "email": student.email,
                    "group": student_group_name.get(student.id, "—"),
                    "telegram_linked": bool(student.telegram_id),
                    "avg_grade": avg_grade,
                    "attendance": attendance,
                }
            )

        rows.sort(key=lambda x: x["full_name"])

    # Generate Excel file
    if not Workbook:
        # Fallback if openpyxl not installed
        raise HTTPException(status_code=500, detail="Excel generation not available")

    wb = Workbook()
    ws = wb.active
    ws.title = "Звіт відвідуваності"

    # Style definitions
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # Headers
    headers = ["Студент", "Email", "Група", "Середня оцінка", "Відвідуваність", "Telegram"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for row_idx, student in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1).value = student["full_name"]
        ws.cell(row=row_idx, column=2).value = student["email"]
        ws.cell(row=row_idx, column=3).value = student["group"]
        ws.cell(row=row_idx, column=4).value = f"{student['avg_grade']}%"
        ws.cell(row=row_idx, column=5).value = f"{student['attendance']}%"
        ws.cell(row=row_idx, column=6).value = "Так" if student["telegram_linked"] else "Ні"

        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"attendance_report_{datetime.date.today()}.xlsx"
    return StreamingResponse(
        iter([excel_file.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/courses")
async def get_teacher_courses(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get list of courses that the teacher teaches"""
    require_teacher_or_admin(current_user)

    # Get teacher's groups
    groups_result = await db.execute(
        select(StudyGroup)
        .where(StudyGroup.teacher_id == current_user.id)
        .options(selectinload(StudyGroup.courses))
    )
    groups = groups_result.scalars().all()

    # Collect all unique course IDs
    course_ids = _collect_course_ids(groups)

    if not course_ids:
        return []

    # Fetch full course details
    from ...models import Course
    courses_result = await db.execute(
        select(Course).where(Course.id.in_(course_ids))
    )
    courses = courses_result.scalars().all()

    return [
        {
            "id": c.id,
            "title": c.title,
            "description": c.description,
        }
        for c in courses
    ]
